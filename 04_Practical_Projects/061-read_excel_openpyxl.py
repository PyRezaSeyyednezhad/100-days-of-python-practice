# -----------------------
# install openpyxl: pip install openpyxl
# -----------------------

# -----------------------
# Create Workbook and Sheet
# -----------------------
def create_workbook():
    from openpyxl import Workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Hello"
    sheet["B1"] = "World"
    sheet["C1"] = "!"
    sheet["A2"] = "Welcome"
    sheet["B2"] = "to"
    sheet["C2"] = "openpyxl"
    sheet["A3"] = "This"
    sheet["B3"] = "is"
    sheet["C3"] = "Excel"
    sheet["D3"] = "File"
    sheet["E3"] = "Created"
    sheet["F3"] = "with"
    sheet["G3"] = "openpyxl"
    sheet["H3"] = "Library"
    workbook.save(filename="./files/openpyxl_files/hello_world.xlsx")
# create_workbook()


# -----------------------
# Read external spreadsheet
# -----------------------
def read_spreadsheet():
    from openpyxl import load_workbook
    workbook = load_workbook(filename="./files/openpyxl_files/openpyxl_sample_data.xlsx")
    print(f"Sheets: {workbook.sheetnames}")
    sheet = workbook.active
    print(f"Active Sheet: {sheet.title}")
    
    # Read cells of sheets
    # First Method: Using []
    print(f"Cell of [H5]: {sheet["H5"].value}")
    print(f"Cell of [A6]: {sheet["A6"].value}")
    print(f"Cell of [C3]: {sheet["C3"].value}")
    
    # Second Method: Using Cell method
    print(f"Cell of [H5](row=5, column=8): {sheet.cell(row=5, column=8).value}")
    print(f"Cell of [A6](row=6, column=1): {sheet.cell(row=6, column=1).value}")
    print(f"Cell of [C3](row=3, column=3): {sheet.cell(row=3, column=3).value}")
# read_spreadsheet()


# -----------------------
# Read_only mode
# -----------------------
def read_only_mode():
    from openpyxl import load_workbook
    workbook = load_workbook(filename="./files/openpyxl_files/openpyxl_sample_data.xlsx", read_only=True)
    print(f"Sheets: {workbook.sheetnames}")
    sheet = workbook.active
    print(f"Active Sheet: {sheet.title}")
    
    # Read cells of sheets
    print(f"Cell of [H5]: {sheet['H5'].value}")
    print(f"Cell of [A6]: {sheet['A6'].value}")
    print(f"Cell of [C3]: {sheet['C3'].value}")
# read_only_mode()


# -----------------------
# Data only mode
# -----------------------
def data_only_mode():
    from openpyxl import load_workbook
    workbook = load_workbook(filename="./files/openpyxl_files/openpyxl_sample_data.xlsx", data_only=True)
    print(f"Sheets: {workbook.sheetnames}")
    sheet = workbook.active
    print(f"Active Sheet: {sheet.title}")
    
    # Read cells of sheets
    print(f"Cell of [H5]: {sheet['H5'].value}")
    print(f"Cell of [A6]: {sheet['A6'].value}")
    print(f"Cell of [C3]: {sheet['C3'].value}")
# data_only_mode()


# -----------------------
# How to access range of cells values
# -----------------------
class access_cell_values:
    def __init__(self): 
        from openpyxl import load_workbook
        workbook = load_workbook(filename="./files/openpyxl_files/openpyxl_sample_data.xlsx")
        self.sheet = workbook.active
    # Method 1: Accessing a range of cells
    def access_with_special_row_column(self):
        print(f"Accessing range of cells A1:C3: {self.sheet['A1:C3']}")
        print("---------------------------------")
        for row in self.sheet["A1:C3"]:
            print(f"Row: {row}")
            for cell in row:
                print(f"Cell of {cell.coordinate}: {cell.value}")
    
    # Method 2: Get All cells from special column (like. column A)
    def access_to_special_column(self):
        print(f"Accessing column A: {self.sheet['A']}")
        print("---------------------------------")
        for cell in self.sheet["A"]:
            print(f"Cell of {cell.coordinate}: {cell.value}")

    # Method 3: Get All cells for range of columns (like. columns A to C)
    def access_to_range_of_columns(self):
        print(f"Accessing columns A to C: {self.sheet['A:C']}")
        print("---------------------------------")
        for column in self.sheet["A:C"]:
            for cell in column:
                print(f"Cell of {cell.coordinate}: {cell.value}")
    
    # Method 4: Get All cells from special row (like. row 5)
    def access_to_special_row(self):
        print(f"Accessing row 5: {self.sheet[5]}")
        for cell in self.sheet[5]:
            print(f"Cell of {cell.coordinate}: {cell.value}")
    
    # Method 5: Get All cells for range of rows (like. rows 1 to 5)
    def access_to_range_of_rows(self):
        print(f"Accessing rows 1 to 5: {self.sheet[1:5]}")
        print("---------------------------------")
        for row in self.sheet[1:5]:
            for cell in row:
                print(f"Cell of {cell.coordinate}: {cell.value}")
    
    # iter_rows() method: Get cells by iterating over rows
    def access_with_iter_rows(self):
        print(f"Accessing cells with iter_rows() method: {self.sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3)}")
        print("---------------------------------")
        for row in self.sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3, values_only=True):
            print(f"Row: {row}")
            for cell in row:
                print(f"{cell}")
        
    
    # iter_cols() method: Get cells by iterating over columns
    def access_with_iter_cols(self):
        print(f"Accessing cells with iter_cols() method: {self.sheet.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3)}")
        print("---------------------------------")
        for column in self.sheet.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3, values_only=True):
            print(f"Column: {column}")
            for cell in column:
                print(f"{cell}")
    
    # Get All data cell info of rows
    def access_all_rows_data(self):
        print(f"Accessing all rows data with values_only=True: {self.sheet.rows}")
        print("---------------------------------")
        for row in self.sheet.rows:
            print(row)

    # Get All data cell info of columns
    def access_all_columns_data(self):
        print(f"Accessing all columns data with values_only=True: {self.sheet.columns}")
        print("---------------------------------")
        for column in self.sheet.columns:
            print(column)


def access_data_of_spreadsheet():
    access = access_cell_values()
    # access.access_with_special_row_column()
    # access.access_to_special_column()
    # access.access_to_range_of_columns()
    # access.access_to_special_row()
    # access.access_to_range_of_rows()
    # access.access_with_iter_rows()
    # access.access_with_iter_cols()
    # access.access_all_rows_data()
    # access.access_all_columns_data()

# access_data_of_spreadsheet()

# HW: Create a function that turn excel header to dictionary
def excel_header_to_dict():
    from openpyxl import load_workbook
    import json
    workbook = load_workbook(filename="./files/openpyxl_files/openpyxl_sample_data.xlsx")
    sheet = workbook.active
    products = {}
    for row in sheet.iter_rows(min_row=2, max_row=5, min_col=4, max_col=7, values_only=True):
        product_id = row[0]
        product = {
            "parent": row[1],
            "title": row[2],
            "category": row[3],
        }
        products[product_id] = product
    
    print(f"Product Header Dictionary: {json.dumps(products)}")

# excel_header_to_dict()

# HW: Convert Excel Header to Python Class
from dataclasses import dataclass
from datetime import datetime
@dataclass
class Product:
    product_id: str
    product_parent: str
    product_title: str
    product_category: str

@dataclass
class Review:
    review_id: str
    review_headline: str
    review_body: str
    star_rating: int
    helpful_votes: int
    total_votes: int
    review_date: datetime

def excel_header_to_class():
    from openpyxl import load_workbook
    workbook = load_workbook(filename="./files/openpyxl_files/openpyxl_sample_data.xlsx")
    sheet = workbook.active
    products = []
    reviews = []
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        product = Product(product_id=row[3], product_parent=row[4], product_title=row[5], product_category=row[6])
        review_date = row[14]
        convert_date = datetime.strptime(review_date, "%Y-%m-%d") if review_date else None
        review = Review(review_id=row[2], review_headline=row[12], review_body=row[13], star_rating=row[7], helpful_votes=row[8], total_votes=row[9], review_date=convert_date)
        
        products.append(product)
        reviews.append(review)
    
    print(f"Products: {products[0]}")
    print(f"Reviews: {reviews[0]}")

# excel_header_to_class()

# insert and remove rows and cols
# insert new rows into spreadsheet
def insert_new_rows_into_spreadsheet(index: int, amount: int):
    from openpyxl import load_workbook
    filename = "./files/openpyxl_files/hello_world.xlsx"
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    sheet.insert_rows(idx=index, amount=amount)
    workbook.save(filename=filename)

# insert_new_rows_into_spreadsheet(1, 1) # Insert one row into the first row
# insert_new_rows_into_spreadsheet(3, 5) # Insert three rows into the third row

# remove rows from spreadsheet
def remove_rows_from_spreadsheet(index: int, amount: int):
    from openpyxl import load_workbook
    filename = "./files/openpyxl_files/hello_world.xlsx"
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    sheet.delete_rows(idx=index, amount=amount)
    workbook.save(filename=filename)

# remove_rows_from_spreadsheet(1,1) # Remove one row from first row
# remove_rows_from_spreadsheet(2, 5) # Remove three rows from second row
    
# insert new cols into spreadsheet
def insert_new_cols_into_spreadsheet(index: int, amount: int):
    from openpyxl import load_workbook
    filename = "./files/openpyxl_files/hello_world.xlsx"
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    sheet.insert_cols(idx=index, amount=amount)
    workbook.save(filename=filename)

# insert_new_cols_into_spreadsheet(1, 1) # Insert one col into the first col
# insert_new_cols_into_spreadsheet(3, 5) # Insert three cols into the third col

# remove cols from spreadsheet
def remove_cols_from_spreadsheet(index: int, amount: int):
    from openpyxl import load_workbook
    filename = "./files/openpyxl_files/hello_world.xlsx"
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    sheet.delete_cols(idx=index, amount=amount)
    workbook.save(filename=filename)

# remove_cols_from_spreadsheet(1,1) # Remove one col from first col
# remove_cols_from_spreadsheet(2, 5) # Remove three cols from second col


# Update the old values with new values
def update_cell_value(cell_pos, new_value):
    from openpyxl import load_workbook
    filename = "./files/openpyxl_files/hello_world.xlsx"
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    sheet[cell_pos] = new_value
    workbook.save(filename=filename)

# update_cell_value("B1", "Everyone") # "World" => "Everyone"

# Working with Sheets
def sheet_manage():
    from openpyxl import load_workbook
    filename = "./files/openpyxl_files/hello_world.xlsx"
    workbook = load_workbook(filename=filename)
    print(f"Sheets: {workbook.sheetnames}")
    sheet = workbook.active
    
    #** Create New Sheets
    # print("Adding new sheets: 'product_sheet' and 'sales_sheet'")
    # product_sheet = workbook.create_sheet('product_sheet', 1)
    # sale_sheet = workbook.create_sheet('sale_sheet', 0)
    # print(f"Sheets: {workbook.sheetnames}")
    
    #** Adding value to created sheets
    # print("Add new values to product_sheet")
    # product_sheet = workbook['product_sheet']
    # product_sheet["A1"] = "This is product sheet and the A1 is Hello"
    # print("Add new values to sale_sheet")
    # sale_sheet = workbook['sale_sheet']
    # sale_sheet["A1"] = "This is sale sheet and the A1 is Hello"
    
    #** Remove sheets
    # print("Remove sale sheet")
    # sale_sheet = workbook['sale_sheet']
    # workbook.remove(sale_sheet)
    # print(f"Sheets: {workbook.sheetnames}")
    workbook.save(filename=filename)
    
sheet_manage()
